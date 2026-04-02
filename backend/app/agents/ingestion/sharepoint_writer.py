"""
SharePoint Excel Writer
Writes data to SharePoint Excel files using Microsoft Graph API
"""
import json
import sys
from pathlib import Path
from typing import List, Dict, Any, Optional
from datetime import datetime
import io

# Fix Unicode encoding on Windows
if sys.stdout.encoding != 'utf-8':
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')

import pandas as pd
import requests
import logging

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("WARNING: openpyxl not installed. Installing now...")
    import subprocess
    result = subprocess.run(
        [sys.executable, "-m", "pip", "install", "openpyxl"],
        capture_output=True,
        text=True
    )
    if result.returncode == 0:
        HAS_OPENPYXL = True
    else:
        print(f"WARNING: Failed to install openpyxl: {result.stderr}")

from .sharepoint_auth import SharePointAuthManager, load_auth_config_from_env

logger = logging.getLogger(__name__)

# Placeholder configuration - to be filled from config.yaml or environment
SHAREPOINT_CONFIG = {
    "unified_solar": {
        "name": "UnifiedSolarData",
        "site_url": "",  # e.g., "https://yourtenant.sharepoint.com/sites/EnergyData"
        "drive_id": "",  # SharePoint document library drive ID
        "file_name": "unified_solar.xlsx",
    },
    "last_7_days": {
        "name": "last_7_days",
        "site_url": "",
        "drive_id": "",
        "file_name": "last_7_days.xlsx",
    },
    "smb_status": {
        "name": "SMB_Status",
        "site_url": "",
        "drive_id": "",
        "file_name": "smb_status.xlsx",
    },
    "grid_and_diesel": {
        "name": "grid_and_diesel",
        "site_url": "",
        "drive_id": "",
        "file_name": "grid_and_diesel.xlsx",
    },
    "master_data": {
        "name": "master_data",
        "site_url": "",
        "drive_id": "",
        "file_name": "master_data.xlsx",
    }
}


class SharePointExcelWriter:
    """
    Writer for Excel files in SharePoint using Microsoft Graph API
    """
    
    def __init__(self, auth_manager: Optional[SharePointAuthManager] = None):
        """
        Initialize SharePoint Excel writer
        
        Args:
            auth_manager: SharePointAuthManager instance (creates new if None)
        """
        if not HAS_OPENPYXL:
            print("ERROR: openpyxl library not available")
            self.auth_manager = None
            self.authenticated = False
            return
        
        if auth_manager:
            self.auth_manager = auth_manager
        else:
            config = load_auth_config_from_env()
            self.auth_manager = SharePointAuthManager(config)
        
        self.authenticated = self.auth_manager.get_access_token() is not None
        self.graph_base_url = "https://graph.microsoft.com/v1.0"
        
        if not self.authenticated:
            print("[WARNING] SharePoint authentication failed. Writer will not function until credentials are provided.")
    
    def _get_or_create_file(self, site_url: str, drive_id: str, file_name: str) -> Optional[str]:
        """
        Get file item ID, creating empty Excel file if it doesn't exist
        
        Args:
            site_url: SharePoint site URL
            drive_id: Drive ID
            file_name: Name of Excel file
            
        Returns:
            Item ID or None
        """
        if not self.authenticated:
            print("[ERROR] Not authenticated with SharePoint")
            return None
        
        try:
            headers = self.auth_manager.get_headers()
            
            # Try to find existing file
            search_url = f"{self.graph_base_url}/drives/{drive_id}/root/children"
            response = requests.get(search_url, headers=headers)
            
            if response.status_code == 200:
                items = response.json().get("value", [])
                for item in items:
                    if item.get("name") == file_name:
                        print(f"[INFO] Found existing file: {file_name} (ID: {item.get('id')})")
                        return item.get("id")
            
            # File doesn't exist, create empty Excel file
            print(f"[INFO] Creating new Excel file: {file_name}")
            
            # Create empty workbook in memory
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            
            # Save to bytes
            excel_bytes = io.BytesIO()
            wb.save(excel_bytes)
            excel_bytes.seek(0)
            
            # Upload to SharePoint
            upload_url = f"{self.graph_base_url}/drives/{drive_id}/root/children/{file_name}/content"
            headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            
            response = requests.put(upload_url, data=excel_bytes.getvalue(), headers=headers)
            
            if response.status_code in [200, 201]:
                item_id = response.json().get("id")
                print(f"[OK] Created new Excel file in SharePoint (ID: {item_id})")
                return item_id
            else:
                print(f"[ERROR] Failed to create file: {response.status_code}")
                print(f"Response: {response.text}")
                return None
        
        except Exception as e:
            print(f"[ERROR] Exception in _get_or_create_file: {str(e)}")
            return None
    
    def _download_excel_file(self, drive_id: str, item_id: str) -> Optional[bytes]:
        """
        Download Excel file content from SharePoint
        
        Args:
            drive_id: Drive ID
            item_id: File item ID
            
        Returns:
            File bytes or None
        """
        if not self.authenticated:
            return None
        
        try:
            headers = self.auth_manager.get_headers()
            download_url = f"{self.graph_base_url}/drives/{drive_id}/items/{item_id}/content"
            response = requests.get(download_url, headers=headers)
            
            if response.status_code == 200:
                return response.content
            else:
                print(f"[ERROR] Failed to download file: {response.status_code}")
                return None
        
        except Exception as e:
            print(f"[ERROR] Failed to download file: {str(e)}")
            return None
    
    def _upload_excel_file(self, drive_id: str, item_id: str, file_bytes: bytes) -> bool:
        """
        Upload Excel file content to SharePoint
        
        Args:
            drive_id: Drive ID
            item_id: File item ID
            file_bytes: Excel file bytes
            
        Returns:
            True if successful
        """
        if not self.authenticated:
            return False
        
        try:
            headers = self.auth_manager.get_headers()
            headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            
            upload_url = f"{self.graph_base_url}/drives/{drive_id}/items/{item_id}/content"
            response = requests.put(upload_url, data=file_bytes, headers=headers)
            
            if response.status_code in [200, 201]:
                print(f"[OK] Successfully uploaded Excel file to SharePoint")
                return True
            else:
                print(f"[ERROR] Failed to upload file: {response.status_code}")
                print(f"Response: {response.text}")
                return False
        
        except Exception as e:
            print(f"[ERROR] Failed to upload file: {str(e)}")
            return False
    
    def append_data_to_sheet(
        self,
        sheet_key: str,
        data: List[Dict[str, Any]],
        sheet_name: str = "Sheet1"
    ) -> bool:
        """
        Append data rows to SharePoint Excel sheet
        
        Args:
            sheet_key: Configuration key (e.g., 'unified_solar')
            data: List of row dictionaries to append
            sheet_name: Name of sheet tab to write to
            
        Returns:
            True if successful
        """
        if not self.authenticated or not HAS_OPENPYXL:
            print("[ERROR] SharePoint not authenticated or openpyxl not available")
            return False
        
        if sheet_key not in SHAREPOINT_CONFIG:
            print(f"[ERROR] Unknown sheet key: {sheet_key}")
            return False
        
        if not data:
            print("[WARNING] No data to append")
            return False
        
        config = SHAREPOINT_CONFIG[sheet_key]
        
        # Check if configuration is complete
        if not config.get("site_url") or not config.get("drive_id"):
            print(f"[WARNING] SharePoint configuration incomplete for {sheet_key}")
            print(f"         Update config.yaml with site_url and drive_id")
            return False
        
        try:
            print(f"\n[INFO] Appending {len(data)} rows to SharePoint: {sheet_key}")
            
            # Get or create file
            item_id = self._get_or_create_file(config["site_url"], config["drive_id"], config["file_name"])
            if not item_id:
                print(f"[ERROR] Could not get/create file: {config['file_name']}")
                return False
            
            # Download existing file
            file_bytes = self._download_excel_file(config["drive_id"], item_id)
            if file_bytes:
                excel_file = io.BytesIO(file_bytes)
                wb = openpyxl.load_workbook(excel_file)
            else:
                # Create new workbook if download fails
                print("[INFO] Creating new workbook (download failed or file empty)")
                from openpyxl import Workbook
                wb = Workbook()
            
            # Get or create sheet
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.create_sheet(sheet_name)
            
            # Add header row if sheet is empty
            if ws.max_row == 1 and not ws[1][0].value:
                headers = list(data[0].keys())
                for col, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col, value=header)
                start_row = 2
            else:
                start_row = ws.max_row + 1
            
            # Append data rows
            headers = list(data[0].keys())
            for row_idx, row_data in enumerate(data, start=start_row):
                for col, header in enumerate(headers, 1):
                    value = row_data.get(header, "")
                    ws.cell(row=row_idx, column=col, value=value)
            
            # Save to bytes
            excel_bytes = io.BytesIO()
            wb.save(excel_bytes)
            excel_bytes.seek(0)
            
            # Upload back to SharePoint
            success = self._upload_excel_file(config["drive_id"], item_id, excel_bytes.getvalue())
            
            if success:
                print(f"[OK] Successfully appended data to SharePoint: {sheet_key}")
                print(f"     ({len(data)} rows added, total rows now: {ws.max_row - 1})")
            
            return success
        
        except Exception as e:
            print(f"[ERROR] Exception appending data to SharePoint: {str(e)}")
            import traceback
            traceback.print_exc()
            return False
    
    def update_last_row(
        self,
        sheet_key: str,
        row_data: Dict[str, Any],
        sheet_name: str = "Sheet1"
    ) -> bool:
        """
        Update the last row in SharePoint Excel sheet
        
        Args:
            sheet_key: Configuration key
            row_data: Dictionary of values to update
            sheet_name: Name of sheet tab
            
        Returns:
            True if successful
        """
        if not self.authenticated or not HAS_OPENPYXL:
            print("[ERROR] SharePoint not authenticated or openpyxl not available")
            return False
        
        if sheet_key not in SHAREPOINT_CONFIG:
            print(f"[ERROR] Unknown sheet key: {sheet_key}")
            return False
        
        config = SHAREPOINT_CONFIG[sheet_key]
        
        if not config.get("site_url") or not config.get("drive_id"):
            print(f"[WARNING] SharePoint configuration incomplete for {sheet_key}")
            return False
        
        try:
            print(f"[INFO] Updating last row in SharePoint: {sheet_key}")
            
            # Get file
            item_id = self._get_or_create_file(config["site_url"], config["drive_id"], config["file_name"])
            if not item_id:
                print(f"[ERROR] Could not find file: {config['file_name']}")
                return False
            
            # Download file
            file_bytes = self._download_excel_file(config["drive_id"], item_id)
            if not file_bytes:
                print("[ERROR] Could not download file")
                return False
            
            # Load workbook
            excel_file = io.BytesIO(file_bytes)
            wb = openpyxl.load_workbook(excel_file)
            
            if sheet_name not in wb.sheetnames:
                print(f"[WARNING] Sheet '{sheet_name}' not found")
                return False
            
            ws = wb[sheet_name]
            
            # Get header row
            headers = {}
            for col in range(1, ws.max_column + 1):
                header_val = ws.cell(row=1, column=col).value
                if header_val:
                    headers[header_val] = col
            
            # Update last row
            last_row = ws.max_row
            for field, value in row_data.items():
                if field in headers:
                    col = headers[field]
                    ws.cell(row=last_row, column=col, value=value)
            
            # Save and upload
            excel_bytes = io.BytesIO()
            wb.save(excel_bytes)
            excel_bytes.seek(0)
            
            success = self._upload_excel_file(config["drive_id"], item_id, excel_bytes.getvalue())
            
            if success:
                print(f"[OK] Successfully updated last row in SharePoint: {sheet_key}")
            
            return success
        
        except Exception as e:
            print(f"[ERROR] Exception updating last row: {str(e)}")
            import traceback
            traceback.print_exc()
            return False


def get_sharepoint_writer() -> SharePointExcelWriter:
    """Get a SharePoint Excel writer instance"""
    config = load_auth_config_from_env()
    auth_manager = SharePointAuthManager(config)
    return SharePointExcelWriter(auth_manager)
